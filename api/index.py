"""
NBPharma IMS — Vercel Serverless Function (Complete Pipeline)
==============================================================
Accepts distributor Excel files via HTTP POST and runs the FULL pipeline:

  MTD Sales Report ->
    1. Enrich ma_imsskubatches (batch-level: PATCH/CREATE)
    2. Upload ma_imsproductdata (metrics: Private, Tender, IMS Total, FOC)

  Closing Stock Report ->
    1. Enrich ma_imsskubatches (batch-level: PATCH/CREATE)
    2. Upload ma_imsproductdata (metrics: Stock Open, Stock Close)

Auth: Refresh token (no App Registration needed).
"""

import os, io, json, time, logging
from datetime import datetime
from collections import defaultdict
import msal
import requests as http_requests
import openpyxl

DATAVERSE_URL = os.environ.get("DATAVERSE_URL", "https://ppcustomprojects.crm4.dynamics.com")
API_URL = f"{DATAVERSE_URL}/api/data/v9.2"
REFRESH_TOKEN = os.environ.get("REFRESH_TOKEN", "")
CLIENT_ID = "51f81489-12ee-4a9e-aaae-a2591f45987d"
AUTHORITY = "https://login.microsoftonline.com/organizations"
API_SECRET_KEY = os.environ.get("API_SECRET_KEY", "nbpharma-ims-poc-2025")
SCOPES = [f"{DATAVERSE_URL}/.default"]

TABLE_SKU_BATCHES = "ma_imsskubatches"
TABLE_PRODUCT_DATA = "ma_imsproductdata"
TABLE_DISTRIBUTORS = "ma_distributors"
TABLE_PRODUCTS = "ma_imsproducts"

UNICARE_PRODUCT_MAP = {
    "U190001": {"nbp_desc": "CIMZIA 200mg/ml 2 Prefilled Syringes ME", "brand": "CIMZIA"},
    "U190003": {"nbp_desc": "Orladeyo 150mg Caps 28's US1", "brand": "ORLADEYO"},
    "U190004": {"nbp_desc": "Livmarli oral solution 30ml bottle, 1's", "brand": "LIVMARLI"},
    "U190005": {"nbp_desc": "Bimzelx Inj 160mg/ml 2 Prefilled Injectors ME1", "brand": "BIMZELX"},
}
MATERIAL_TO_BRAND = {k: v["brand"] for k, v in UNICARE_PRODUCT_MAP.items()}

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

# ── AUTH ──
_token_cache = {"token": None, "expires_at": 0}

def get_token():
    now = time.time()
    if _token_cache["token"] and _token_cache["expires_at"] > now + 60:
        return _token_cache["token"]
    if not REFRESH_TOKEN:
        raise Exception("REFRESH_TOKEN env var not set.")
    app_client = msal.PublicClientApplication(client_id=CLIENT_ID, authority=AUTHORITY)
    result = app_client.acquire_token_by_refresh_token(refresh_token=REFRESH_TOKEN, scopes=SCOPES)
    if "access_token" in result:
        _token_cache["token"] = result["access_token"]
        _token_cache["expires_at"] = now + result.get("expires_in", 3600)
        return result["access_token"]
    raise Exception(f"Auth failed: {result.get('error_description', 'Unknown')}")

def get_headers():
    return {"Authorization": f"Bearer {get_token()}", "Content-Type": "application/json",
            "OData-MaxVersion": "4.0", "OData-Version": "4.0", "Accept": "application/json"}

# ── DATAVERSE HELPERS ──
_odata_cache = {}

def resolve_odata_set(logical_name):
    if logical_name in _odata_cache:
        return _odata_cache[logical_name]
    for suffix in ["s", "es", "set", ""]:
        candidate = f"{logical_name}{suffix}"
        resp = http_requests.get(f"{API_URL}/{candidate}?$top=1", headers=get_headers(), timeout=10)
        if resp.status_code == 200:
            _odata_cache[logical_name] = candidate
            return candidate
    return f"{logical_name}s"

def discover_valid_columns(table_name):
    url = (f"{API_URL}/EntityDefinitions(LogicalName='{table_name}')"
           f"/Attributes?$select=LogicalName,AttributeType"
           f"&$filter=AttributeType ne Microsoft.Dynamics.CRM.AttributeTypeCode'Virtual'")
    resp = http_requests.get(url, headers=get_headers(), timeout=15)
    columns = set()
    primary_name = None
    if resp.status_code == 200:
        for a in resp.json().get("value", []):
            ln = a["LogicalName"]
            if ln.startswith("ma_"):
                columns.add(ln)
                if a["AttributeType"] == "String" and not primary_name:
                    if "batchid" in ln or "dataid" in ln or "name" in ln or "transactionid" in ln:
                        primary_name = ln
    return columns, primary_name

def discover_metric_picklist():
    url = (f"{API_URL}/EntityDefinitions(LogicalName='{TABLE_PRODUCT_DATA}')"
           f"/Attributes/Microsoft.Dynamics.CRM.PicklistAttributeMetadata"
           f"?$filter=LogicalName eq 'ma_metric'&$expand=OptionSet")
    resp = http_requests.get(url, headers=get_headers(), timeout=15)
    metric_map = {}
    if resp.status_code == 200:
        data = resp.json().get("value", [])
        if data:
            for opt in data[0].get("OptionSet", {}).get("Options", []):
                lbl = (opt.get("Label", {}).get("UserLocalizedLabel") or {}).get("Label", "")
                metric_map[lbl.upper()] = opt["Value"]
                metric_map[opt["Value"]] = lbl
    return metric_map

def safe_payload(payload, valid_columns):
    clean = {}
    for k, v in payload.items():
        if "@odata.bind" in k:
            clean[k] = v
        elif k.startswith("ma_") and k in valid_columns:
            clean[k] = v
    return clean

def find_product_guid(brand_name, all_products):
    bl = brand_name.lower()
    for p in all_products:
        for k, v in p.items():
            if k.startswith("ma_") and isinstance(v, str) and bl in v.lower():
                return p["ma_imsproductsid"]
    return None


def match_product_dynamically(product_desc, all_products):
    """Match a product description from file against ma_imsproducts.
    Uses keyword overlap scoring — works across different distributor naming conventions."""
    if not product_desc:
        return None, None

    desc_upper = product_desc.upper()
    desc_words = set(desc_upper.replace(",", " ").replace("-", " ").replace("'", "").split())
    # Remove very short/common words
    desc_words = {w for w in desc_words if len(w) >= 3 and w not in {"THE", "FOR", "AND", "MG/ML"}}

    best_guid = None
    best_name = None
    best_score = 0

    for p in all_products:
        for field in ["ma_sku", "ma_productid", "ma_productname", "ma_name"]:
            pname = p.get(field)
            if not pname or not isinstance(pname, str):
                continue
            pname_upper = pname.upper()
            pname_words = set(pname_upper.replace(",", " ").replace("-", " ").replace("'", "").split())
            pname_words = {w for w in pname_words if len(w) >= 3}

            if not pname_words:
                continue

            overlap = desc_words & pname_words
            if not overlap:
                continue

            # Score: overlap ratio + bonus for first word (brand name) match
            score = len(overlap) / max(len(desc_words | pname_words), 1)
            # Big bonus if the first significant word matches (brand name)
            first_word = desc_upper.split()[0] if desc_upper.split() else ""
            if first_word in pname_upper:
                score += 0.5

            if score > best_score:
                best_score = score
                best_guid = p["ma_imsproductsid"]
                best_name = pname

    if best_score >= 0.3:  # threshold
        return best_guid, best_name
    return None, None


DISTRIBUTOR_CODES = {
    "UNI": "Unicare",
    "MPC": "Modern Pharmaceutical",
    "PTD": "Pharmatrade",
    "PHL": "Pharmalink",
    "EBS": "Ebn Sina",
}


def parse_distributor_from_subject(subject):
    """Extract distributor from email subject.
    Supports: short codes (PTD, UNI, MPC) or full names.
    Examples: 'MTD Sales PTD Nov 2025', 'Closing Stock - Unicare - Sep 2025'"""
    if not subject:
        return None
    subject_upper = subject.upper()

    # Check short codes first
    for code, name in DISTRIBUTOR_CODES.items():
        if code in subject_upper.split():
            return name
        # Also check with dashes: "MTD-PTD-Nov"
        if f"-{code}-" in subject_upper or f"-{code} " in subject_upper or f" {code}-" in subject_upper:
            return name

    # Check full names
    for code, name in DISTRIBUTOR_CODES.items():
        if name.upper() in subject_upper:
            return name

    # Fallback: try splitting by delimiters
    parts = [p.strip() for p in subject.replace("–", "-").split("-")]
    skip_words = {"mtd", "sales", "report", "closing", "stock", "monthly", "product",
                  "january", "february", "march", "april", "may", "june", "july",
                  "august", "september", "october", "november", "december",
                  "2024", "2025", "2026", "2027"}
    for part in parts:
        words = part.lower().split()
        if words and not all(w in skip_words for w in words):
            if len(part) > 2 and part.lower() not in skip_words:
                return part

    return None


def lookup_distributor(name_hint):
    """Lookup distributor by name. Returns (guid, name) or (None, None)."""
    dist_set = resolve_odata_set(TABLE_DISTRIBUTORS)

    if name_hint:
        # Try exact search first
        params = {"$filter": f"contains(ma_distributorname,'{name_hint}')", "$top": 5}
        resp = http_requests.get(f"{API_URL}/{dist_set}", headers=get_headers(), params=params, timeout=15)
        if resp.status_code == 200 and resp.json().get("value"):
            rec = resp.json()["value"][0]
            return rec["ma_distributorsid"], rec["ma_distributorname"]

    # Fallback: try "Unicare"
    params = {"$filter": "contains(ma_distributorname,'Unicare')", "$top": 5}
    resp = http_requests.get(f"{API_URL}/{dist_set}", headers=get_headers(), params=params, timeout=15)
    if resp.status_code == 200 and resp.json().get("value"):
        rec = resp.json()["value"][0]
        return rec["ma_distributorsid"], rec["ma_distributorname"]

    return None, None


def parse_expiry_string(val):
    """Parse expiry date from various formats to YYYY-MM-DD."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try dd.mm.yyyy
    if "." in s:
        try:
            parts = s.split(".")
            if len(parts) == 3:
                return f"{parts[2]}-{parts[1]}-{parts[0]}"
        except (IndexError, ValueError):
            pass
    # Try yyyy-mm-dd (already correct)
    if len(s) >= 10 and s[4] == "-":
        return s[:10]
    return None

# ── DYNAMIC COLUMN MAPPER ──
# Auto-detects column positions from headers — works with ANY distributor template

COLUMN_RULES = {
    "material_code": ["material", "item number", "item no", "item code", "sku", "product code", "matl"],
    "product_name": ["material name", "product description", "item description", "item name", "description", "product name"],
    "customer_name": ["ship to name", "customer name", "customer", "sold to name", "sold-to party", "ship-to party"],
    "customer_code": ["ship to", "customer id", "customer code", "sold to", "sold-to party"],
    "customer_group": ["customer group", "cust group", "channel", "customer group 1"],
    "region": ["region", "city", "area", "territory", "sales district"],
    "salesman": ["salesman", "sales rep", "rep", "sales office"],
    "batch": ["batch", "lot", "lot number", "batch number", "batch no", "batch no."],
    "expiry": ["expiry", "exp date", "expiry date", "batch expiry", "best before", "expire date"],
    "qty": ["nett sales quantity", "normal qty", "salesqty", "total qty", "total batch qty on hand",
            "quantity", "sales qty", "base unit quantity", "closing stock"],
    "value": ["nett sales value", "net value", "salesvalue", "value", "amount", "sales value"],
    "foc_qty": ["foc qty", "foc quantity", "free goods", "focqty", "sum foc"],
    "sample_qty": ["sample qty", "sample quantity"],
    "bill_type": ["billt", "bill. doc. type", "billing type", "doc type"],
    "bill_date": ["bill date", "billing effective date", "invoice date", "billing date", "calendar day"],
    "sales_doc": ["sales doc", "sales order", "invoice number", "document number", "billing document"],
    "bill_doc": ["bill.docs", "billing document", "invoice number"],
    "po_number": ["purchase order", "po number", "cust. reference", "reference"],
    "brand": ["brand", "principal", "manufacturer", "supplier"],
    "opening": ["opening stock", "opening bal", "opening", "sum private stock"],
    "closing": ["closing stock", "closing bal", "closing", "total batch qty on hand"],
    "from_date": ["from date", "start date", "period start"],
    "to_date": ["to date", "end date", "period end"],
    "pack_size": ["pack size", "pack", "bun", "uom", "inventory uom"],
    "unit_price": ["unit price", "price"],
    "mfg_date": ["manufacture date", "mfg date", "manufacturing date"],
    "period": ["cal. year / month", "period", "month/year"],
    "private_qty": ["sum private stock"],
    "tender_qty": ["sum tender stock"],
}


def detect_header_row(ws, max_scan=25):
    """Auto-detect which row contains the actual data headers.
    Returns (header_row_number, headers_list). 1-indexed.
    Heuristic: the row with the most non-empty cells that look like headers."""
    best_row = 1
    best_score = 0
    best_headers = []

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(v or "").strip() for v in row]
        non_empty = [c for c in cells if c and c != "None"]
        if len(non_empty) < 3:
            continue

        # Score: more non-empty cells = more likely to be header row
        score = len(non_empty)
        # Bonus for known header keywords
        header_keywords = {"batch", "material", "item", "customer", "qty", "value", "stock",
                          "date", "product", "sales", "code", "name", "region", "quantity"}
        keyword_hits = sum(1 for c in non_empty if any(kw in c.lower() for kw in header_keywords))
        score += keyword_hits * 2

        if score > best_score:
            best_score = score
            best_row = i
            best_headers = cells

    return best_row, best_headers


def auto_map_columns(headers):
    """Auto-detect column positions from headers. Returns {field_key: col_index}.
    Uses two-pass matching: exact match first, then substring match."""
    mapping = {}
    headers_lower = [(h or "").strip().lower() for h in headers]

    # Pass 1: exact matches only
    for field_key, keywords in COLUMN_RULES.items():
        for kw in keywords:
            for idx, h in enumerate(headers_lower):
                if h == kw:
                    if field_key not in mapping:
                        mapping[field_key] = idx
                        break
            if field_key in mapping:
                break

    # Pass 2: substring matches for remaining unmapped fields
    for field_key, keywords in COLUMN_RULES.items():
        if field_key in mapping:
            continue
        for kw in keywords:
            if len(kw) < 4:
                continue  # Skip short keywords for substring matching
            for idx, h in enumerate(headers_lower):
                if kw in h and idx not in mapping.values():
                    mapping[field_key] = idx
                    break
            if field_key in mapping:
                break

    return mapping


def get_col(row, mapping, field_key, default=None):
    """Safely get a column value by mapped field key."""
    idx = mapping.get(field_key)
    if idx is not None and idx < len(row):
        val = row[idx]
        return val if val is not None else default
    return default


def get_col_str(row, mapping, field_key, default=""):
    val = get_col(row, mapping, field_key, default)
    return str(val).strip() if val is not None else default


def get_col_float(row, mapping, field_key, default=0):
    val = get_col(row, mapping, field_key, default)
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def detect_month_year(row, mapping):
    """Extract month/year from date fields or period fields."""
    # Try period field first (e.g., "11.2025" or "Sep-2025")
    period_val = get_col(row, mapping, "period")
    if period_val:
        s = str(period_val).strip()
        # Format: mm.yyyy or mm/yyyy
        if "." in s or "/" in s:
            parts = s.replace("/", ".").split(".")
            try:
                if len(parts) == 2:
                    return int(parts[0]), int(parts[1])
            except (ValueError, IndexError):
                pass
        # Format: Mon-YYYY (e.g., "Sep-2025")
        month_names = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                       "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        for mn, mv in month_names.items():
            if mn in s.lower():
                for part in s.split("-"):
                    try:
                        yr = int(part)
                        if yr > 2000:
                            return mv, yr
                    except ValueError:
                        continue

    # Try date fields
    for date_field in ["bill_date", "from_date"]:
        val = get_col(row, mapping, date_field)
        if val:
            if isinstance(val, datetime):
                return val.month, val.year
            s = str(val).strip()
            if "." in s:
                try:
                    parts = s.split(".")
                    if len(parts) == 3:
                        return int(parts[1]), int(parts[2])
                except (IndexError, ValueError):
                    pass
            # Try mm/dd/yyyy
            if "/" in s:
                try:
                    parts = s.split("/")
                    if len(parts) == 3:
                        return int(parts[0]), int(parts[2])
                except (IndexError, ValueError):
                    pass
    return None, None


# ── PARSERS (DYNAMIC) ──

def parse_closing_stock(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    header_row, headers = detect_header_row(ws)
    col = auto_map_columns(headers)
    log.info(f"Closing Stock: header row={header_row}, auto-mapped={col}")

    batches = {}
    product_agg = defaultdict(lambda: {"opening": 0, "closing": 0, "description": "", "month": 9, "year": 2025})
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        material = get_col_str(row, col, "material_code")
        if not material:
            continue
        batch_nb = get_col_str(row, col, "batch")
        month, year = detect_month_year(row, col)
        if not month:
            month, year = 9, 2025

        expiry = get_col(row, col, "expiry")
        expiry_str = expiry.strftime("%Y-%m-%d") if isinstance(expiry, datetime) else (str(expiry)[:10] if expiry else None)
        opening = int(get_col_float(row, col, "opening", 0))
        closing = int(get_col_float(row, col, "closing", 0))

        if batch_nb:
            batches[batch_nb] = {"material": material, "description": get_col_str(row, col, "product_name"),
                "opening": opening, "closing": closing, "expiry": expiry_str,
                "month": month, "year": year, "nbp_info": UNICARE_PRODUCT_MAP.get(material, {})}
        product_agg[material]["opening"] += opening
        product_agg[material]["closing"] += closing
        product_agg[material]["description"] = get_col_str(row, col, "product_name")
        product_agg[material]["month"] = month
        product_agg[material]["year"] = year
    wb.close()
    return batches, dict(product_agg)

def parse_mtd_sales(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    # Try active sheet first, then look for "Table" sheet (SAP BEx reports)
    ws = wb.active
    header_row, headers = detect_header_row(ws)

    # If active sheet has very few data columns, try "Table" sheet
    non_empty_headers = [h for h in headers if h.strip()]
    if len(non_empty_headers) < 5 and "Table" in wb.sheetnames:
        ws = wb["Table"]
        header_row, headers = detect_header_row(ws)
        log.info(f"Switched to 'Table' sheet for data")

    col = auto_map_columns(headers)

    # For SAP BEx: unnamed columns next to "Material" often contain product description
    if "material_code" in col and "product_name" not in col:
        mc_idx = col["material_code"]
        # Check if the next column has product-like text
        if mc_idx + 1 < len(headers):
            col["product_name"] = mc_idx + 1

    log.info(f"MTD Sales: header row={header_row}, auto-mapped={col}")

    batches = defaultdict(lambda: {"qty": 0, "value": 0, "material": "", "name": "", "brand": "", "expiry": None})
    product_metrics = defaultdict(lambda: {"private_qty": 0, "private_val": 0, "tender_qty": 0, "tender_val": 0,
                                           "foc_qty": 0, "foc_val": 0, "total_qty": 0, "total_val": 0})
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Get product identifier — try brand first, then material name, then material code
        brand = get_col_str(row, col, "brand")
        product_name = get_col_str(row, col, "product_name")
        material_code = get_col_str(row, col, "material_code")

        # Derive brand from product name if no brand column
        if not brand and product_name:
            brand = product_name.split()[0].upper() if product_name.split() else ""
        if not brand and material_code:
            # Try UNICARE_PRODUCT_MAP
            info = UNICARE_PRODUCT_MAP.get(material_code, {})
            brand = info.get("brand", material_code)
        if not brand:
            continue

        batch = get_col_str(row, col, "batch")
        custgroup = get_col_str(row, col, "customer_group").lower()
        billt = get_col_str(row, col, "bill_type")

        qty = get_col_float(row, col, "qty", 0)
        val = get_col_float(row, col, "value", 0)
        foc = get_col_float(row, col, "foc_qty", 0)

        # Batch-level aggregation
        if batch:
            batches[batch]["qty"] += qty
            batches[batch]["value"] += val
            batches[batch]["material"] = material_code
            batches[batch]["name"] = product_name
            batches[batch]["brand"] = brand
            exp = get_col(row, col, "expiry")
            if exp and isinstance(exp, datetime):
                batches[batch]["expiry"] = exp.strftime("%Y-%m-%d")
            elif exp and isinstance(exp, str):
                try:
                    parts = exp.strip().split(".")
                    if len(parts) == 3:
                        batches[batch]["expiry"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                except (IndexError, ValueError):
                    pass
            # Detect month/year from transaction date
            m, y = detect_month_year(row, col)
            if m and y:
                batches[batch]["month"] = m
                batches[batch]["year"] = y

        # FOC detection: explicit FOC column OR billing type
        is_foc = foc > 0 or billt in ("ZRE5", "ZUFA")

        if is_foc:
            product_metrics[brand]["foc_qty"] += abs(foc) if foc > 0 else abs(qty)
            product_metrics[brand]["foc_val"] += abs(val)
            continue

        # Regular sales by customer group
        if custgroup in ("private", ""):
            product_metrics[brand]["private_qty"] += qty
            product_metrics[brand]["private_val"] += val
        elif custgroup == "institutional":
            product_metrics[brand]["tender_qty"] += qty
            product_metrics[brand]["tender_val"] += val
        else:
            product_metrics[brand]["private_qty"] += qty
            product_metrics[brand]["private_val"] += val

        product_metrics[brand]["total_qty"] += qty
        product_metrics[brand]["total_val"] += val
        # Store month/year from first transaction
        if "month" not in product_metrics[brand] or not product_metrics[brand].get("month"):
            m, y = detect_month_year(row, col)
            if m and y:
                product_metrics[brand]["month"] = m
                product_metrics[brand]["year"] = y
    wb.close()
    return dict(batches), dict(product_metrics)

def detect_file_type(filename):
    fn = filename.lower()
    if "closing" in fn or "stock" in fn:
        return "closing_stock"
    elif "mtd" in fn or "sales" in fn:
        return "mtd_sales"
    return "unknown"

# ── PIPELINE A: ENRICH BATCHES ──

def pipeline_enrich_batches(stock_batches, mtd_batches, dist_guid, brand_guids,
                            batch_set, dist_set, prod_set, valid_columns, primary_name, all_products=None):
    results = {"patched": 0, "created": 0, "failed": 0, "errors": []}
    if all_products is None:
        all_products = []
    all_batch_numbers = set(stock_batches.keys()) | set(mtd_batches.keys())
    if not all_batch_numbers:
        return results
    found_batches = {}
    missing_batches = []
    for batch_nb in sorted(all_batch_numbers):
        url = f"{API_URL}/{batch_set}?$filter=ma_batchnb eq '{batch_nb}'&$top=100"
        resp = http_requests.get(url, headers=get_headers(), timeout=15)
        if resp.status_code == 200:
            records = resp.json().get("value", [])
            if records:
                found_batches[batch_nb] = records
            else:
                missing_batches.append(batch_nb)
        time.sleep(0.05)

    def build_batch_payload(batch_nb):
        payload = {"ma_Distributor@odata.bind": f"/{dist_set}({dist_guid})"}
        if batch_nb in stock_batches:
            sd = stock_batches[batch_nb]
            payload.update({"ma_month": sd.get("month", 9), "ma_year": sd.get("year", 2025),
                "ma_quantity": max(0, sd["closing"]), "ma_openingqty": max(0, sd["opening"]), "ma_closingqty": max(0, sd["closing"])})
            nbp_info = sd.get("nbp_info", {})
            if nbp_info:
                payload["ma_itemno"] = nbp_info["nbp_desc"]
            if sd.get("expiry"):
                exp = parse_expiry_string(sd["expiry"])
                if exp:
                    payload["ma_expirydate"] = exp
            brand = nbp_info.get("brand", "")
            if brand in brand_guids:
                payload["ma_Product@odata.bind"] = f"/{prod_set}({brand_guids[brand]})"
        elif batch_nb in mtd_batches:
            md = mtd_batches[batch_nb]
            # Detect month/year from batch data dates or default
            b_month = md.get("month", 0)
            b_year = md.get("year", 0)
            if not b_month or not b_year:
                b_month, b_year = 9, 2025  # fallback
            payload.update({"ma_month": b_month, "ma_year": b_year, "ma_quantity": max(0, int(md["qty"])), "ma_itemno": md["name"]})
            if md.get("expiry"):
                exp = parse_expiry_string(md["expiry"])
                if exp:
                    payload["ma_expirydate"] = exp
            brand = md.get("brand", "")
            if brand in brand_guids:
                payload["ma_Product@odata.bind"] = f"/{prod_set}({brand_guids[brand]})"
            elif md.get("name"):
                # Try dynamic match by product description
                guid, _ = match_product_dynamically(md["name"], all_products)
                if guid:
                    payload["ma_Product@odata.bind"] = f"/{prod_set}({guid})"
        return payload

    for batch_nb, records in found_batches.items():
        payload = safe_payload(build_batch_payload(batch_nb), valid_columns)
        for rec in records:
            resp = http_requests.patch(f"{API_URL}/{batch_set}({rec['ma_imsskubatchesid']})",
                headers=get_headers(), json=payload, timeout=30)
            if resp.status_code in (200, 204):
                results["patched"] += 1
            else:
                results["failed"] += 1
                results["errors"].append(f"PATCH {batch_nb}: {resp.status_code} {resp.text[:300]}")
            time.sleep(0.1)

    for batch_nb in missing_batches:
        payload = build_batch_payload(batch_nb)
        payload["ma_batchnb"] = batch_nb
        brand = ""
        if batch_nb in stock_batches:
            brand = stock_batches[batch_nb].get("nbp_info", {}).get("brand", "")
        elif batch_nb in mtd_batches:
            brand = mtd_batches[batch_nb].get("brand", "")
        if primary_name:
            yr = payload.get("ma_year", 2025)
            mn = payload.get("ma_month", 9)
            payload[primary_name] = f"{brand}_{batch_nb}_UNI_{yr}-{mn:02d}"
        payload = safe_payload(payload, valid_columns)
        resp = http_requests.post(f"{API_URL}/{batch_set}", headers=get_headers(), json=payload, timeout=30)
        if resp.status_code in (200, 201, 204):
            results["created"] += 1
        else:
            results["failed"] += 1
            results["errors"].append(f"CREATE {batch_nb}: {resp.status_code}")
        time.sleep(0.1)
    return results

# ── PIPELINE B: UPLOAD PRODUCT METRICS (with dedup) ──

def pipeline_upload_metrics(file_type, stock_product_agg, mtd_product_metrics,
                            dist_guid, brand_guids, dist_set, prod_set):
    results = {"metrics_created": 0, "metrics_updated": 0, "metrics_failed": 0, "errors": []}
    metric_map = discover_metric_picklist()
    if not metric_map:
        results["errors"].append("Could not discover metric picklist")
        return results
    pd_columns, pd_primary = discover_valid_columns(TABLE_PRODUCT_DATA)
    if not pd_primary:
        pd_primary = "ma_dataid"
    pd_set = resolve_odata_set(TABLE_PRODUCT_DATA)

    METRICS = {}
    for label_upper, val in metric_map.items():
        if not isinstance(label_upper, str):
            continue
        if "DISTRIBUTORS OPENING STOCK" in label_upper:
            METRICS["stock_open"] = val
        elif "REPORTED CLOSING STOCK" in label_upper:
            METRICS["stock_close"] = val
        elif "PRIVATE MARKET SALES" in label_upper:
            METRICS["private_sales"] = val
        elif "SUPPLY TO LOCAL TENDERS" in label_upper:
            METRICS["tender_sales"] = val
        elif "IMS ACT" in label_upper and "QTTY" in label_upper:
            METRICS["ims_total_qty"] = val
        elif "FOC" in label_upper and "SAMPLE" in label_upper:
            METRICS["foc_samples"] = val
        elif "TOTAL NON SALES" in label_upper:
            METRICS["total_non_sales"] = val

    log.info(f"Mapped {len(METRICS)} metric keys: {list(METRICS.keys())}")
    records = []

    def add_metric(brand, metric_key, value, month, year):
        if metric_key not in METRICS or brand not in brand_guids or value == 0:
            return
        records.append({
            "ma_Distributor@odata.bind": f"/{dist_set}({dist_guid})",
            "ma_Product@odata.bind": f"/{prod_set}({brand_guids[brand]})",
            "ma_metric": METRICS[metric_key], "ma_month": month, "ma_year": year,
            "ma_value": int(round(value)),
            pd_primary: f"{metric_key}_{brand}_UNI_{year}-{month:02d}",
            "_product_guid": brand_guids[brand],
            "_metric_int": METRICS[metric_key],
        })

    if file_type == "closing_stock" and stock_product_agg:
        for material, vals in stock_product_agg.items():
            brand = MATERIAL_TO_BRAND.get(material)
            if not brand:
                continue
            month = vals.get("month", 9)
            year = vals.get("year", 2025)
            add_metric(brand, "stock_open", vals["opening"], month, year)
            add_metric(brand, "stock_close", vals["closing"], month, year)

    if file_type == "mtd_sales" and mtd_product_metrics:
        # Detect month/year from the data (use first batch's date or fallback)
        detected_month, detected_year = 9, 2025
        # Check if parse returned month/year info
        for brand, v in mtd_product_metrics.items():
            if v.get("month") and v.get("year"):
                detected_month = v["month"]
                detected_year = v["year"]
                break
        for brand, v in mtd_product_metrics.items():
            add_metric(brand, "private_sales", v["private_qty"], detected_month, detected_year)
            add_metric(brand, "tender_sales", v["tender_qty"], detected_month, detected_year)
            add_metric(brand, "ims_total_qty", v["total_qty"], detected_month, detected_year)
            add_metric(brand, "foc_samples", v["foc_qty"], detected_month, detected_year)
            add_metric(brand, "total_non_sales", v["foc_qty"], detected_month, detected_year)

    log.info(f"Prepared {len(records)} metric records for ma_imsproductdata")

    # DEDUP: check if record exists before creating
    for rec in records:
        product_guid = rec.pop("_product_guid")
        metric_int = rec.pop("_metric_int")
        month = rec["ma_month"]
        year = rec["ma_year"]

        # Search for existing record with same product + metric + month + year
        filter_str = (f"_ma_product_value eq {product_guid}"
                      f" and ma_metric eq {metric_int}"
                      f" and ma_month eq {month}"
                      f" and ma_year eq {year}"
                      f" and _ma_distributor_value eq {dist_guid}")
        check_url = f"{API_URL}/{pd_set}?$filter={filter_str}&$top=1"
        check_resp = http_requests.get(check_url, headers=get_headers(), timeout=15)

        existing = None
        if check_resp.status_code == 200:
            existing_records = check_resp.json().get("value", [])
            if existing_records:
                existing = existing_records[0]

        clean = safe_payload(rec, pd_columns)

        if existing:
            # PATCH existing record (update value)
            rec_id = existing.get("ma_imsproductdataid") or existing.get("ma_imsproductdatasid")
            if not rec_id:
                for k, v in existing.items():
                    if k.endswith("id") and k.startswith("ma_") and v:
                        rec_id = v
                        break
            if rec_id:
                patch_payload = {"ma_value": clean.get("ma_value", 0)}
                resp = http_requests.patch(f"{API_URL}/{pd_set}({rec_id})",
                    headers=get_headers(), json=patch_payload, timeout=30)
                if resp.status_code in (200, 204):
                    results["metrics_updated"] += 1
                else:
                    results["metrics_failed"] += 1
                    results["errors"].append(f"METRIC PATCH: {resp.status_code}")
            else:
                results["metrics_failed"] += 1
                results["errors"].append("Could not find record ID for existing metric")
        else:
            # CREATE new record
            resp = http_requests.post(f"{API_URL}/{pd_set}", headers=get_headers(), json=clean, timeout=30)
            if resp.status_code in (200, 201, 204):
                results["metrics_created"] += 1
            else:
                results["metrics_failed"] += 1
                results["errors"].append(f"METRIC CREATE: {resp.status_code}")
        time.sleep(0.1)
    return results

# ── PIPELINE C: UPLOAD RAW TRANSACTIONS (ma_IMSRawTransactions) ──

def pipeline_upload_raw_transactions(file_type, file_bytes, dist_guid, brand_guids, dist_set, prod_set):
    """Upload individual transaction rows from MTD Sales to ma_IMSRawTransactions."""
    results = {"raw_created": 0, "raw_failed": 0, "errors": []}

    if file_type != "mtd_sales":
        return results

    try:
        raw_set = resolve_odata_set("ma_imsrawtransactions")
    except Exception:
        results["errors"].append("ma_imsrawtransactions table not accessible, skipping")
        return results

    raw_columns, raw_primary = discover_valid_columns("ma_imsrawtransactions")
    if not raw_columns:
        results["errors"].append("Could not discover raw transactions columns")
        return results

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    # Try "Table" sheet for SAP BEx reports
    if "Table" in wb.sheetnames:
        ws = wb["Table"]
    header_row, headers = detect_header_row(ws)
    col = auto_map_columns(headers)
    # SAP BEx: product description in unnamed column next to material
    if "material_code" in col and "product_name" not in col:
        mc_idx = col["material_code"]
        if mc_idx + 1 < len(headers):
            col["product_name"] = mc_idx + 1
    log.info(f"Raw TX: header row={header_row}, auto-mapped: {col}")
    row_count = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        product_name = get_col_str(row, col, "product_name")
        if not product_name:
            continue
        row_count += 1

        # Derive brand
        brand = get_col_str(row, col, "brand")
        if not brand and product_name:
            brand = product_name.split()[0].upper()

        bill_date = get_col(row, col, "bill_date")
        date_str = bill_date.strftime("%Y-%m-%d") if isinstance(bill_date, datetime) else None
        month, year = detect_month_year(row, col)
        if not month:
            month, year = 9, 2025

        payload = {"ma_distributor@odata.bind": f"/{dist_set}({dist_guid})"}
        if brand in brand_guids:
            payload["ma_product@odata.bind"] = f"/{prod_set}({brand_guids[brand]})"

        field_map = {
            "ma_distributoritemcode": get_col_str(row, col, "material_code"),
            "ma_itemdescription": product_name,
            "ma_customername": get_col_str(row, col, "customer_name"),
            "ma_region": get_col_str(row, col, "region"),
            "ma_batchnumber": get_col_str(row, col, "batch"),
            "ma_billingdocument": get_col_str(row, col, "bill_doc") or get_col_str(row, col, "sales_doc"),
            "ma_ponumber": get_col_str(row, col, "po_number"),
            "ma_sourcefile": "auto_import",
        }
        for k, v in field_map.items():
            if k in raw_columns and v:
                payload[k] = v[:200] if len(v) > 200 else v

        # Numeric fields — ma_quantity and ma_value don't exist on this table
        qty = get_col_float(row, col, "qty", 0)
        val = get_col_float(row, col, "value", 0)
        foc = get_col_float(row, col, "foc_qty", 0)

        if "ma_mappedsku" in raw_columns:
            payload["ma_mappedsku"] = f"qty={qty:.0f}|val={val:.2f}"

        # Integer fields
        if "ma_month" in raw_columns:
            payload["ma_month"] = month
        if "ma_year" in raw_columns:
            payload["ma_year"] = year

        # Date fields
        if date_str and "ma_transactiondate" in raw_columns:
            payload["ma_transactiondate"] = date_str
        exp = get_col(row, col, "expiry")
        if exp and isinstance(exp, datetime) and "ma_expirydate" in raw_columns:
            payload["ma_expirydate"] = exp.strftime("%Y-%m-%d")
        elif exp and isinstance(exp, str) and "ma_expirydate" in raw_columns:
            # Try parsing dd.mm.yyyy
            try:
                parts = exp.strip().split(".")
                if len(parts) == 3:
                    payload["ma_expirydate"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
            except (IndexError, ValueError):
                pass

        # Boolean: FOC flag
        billt = get_col_str(row, col, "bill_type")
        if "ma_isfoc" in raw_columns:
            payload["ma_isfoc"] = foc > 0 or billt in ("ZRE5", "ZUFA")

        # Primary name: transactionid (used for dedup)
        if raw_primary:
            sales_doc = get_col_str(row, col, "sales_doc") or str(row_count)
            cust_code = get_col_str(row, col, "customer_code") or ""
            tx_id = f"TX-{sales_doc}-{cust_code}-{year}-{month:02d}"
            payload[raw_primary] = tx_id

        clean = safe_payload(payload, raw_columns)

        # DEDUP: check if this transaction already exists
        if raw_primary and tx_id:
            check_filter = f"{raw_primary} eq '{tx_id}'"
            check_url = f"{API_URL}/{raw_set}?$filter={check_filter}&$top=1&$select={raw_primary}"
            check_resp = http_requests.get(check_url, headers=get_headers(), timeout=10)
            if check_resp.status_code == 200 and check_resp.json().get("value"):
                results.setdefault("raw_skipped", 0)
                results["raw_skipped"] += 1
                continue  # Already exists, skip

        resp = http_requests.post(f"{API_URL}/{raw_set}", headers=get_headers(), json=clean, timeout=30)
        if resp.status_code in (200, 201, 204):
            results["raw_created"] += 1
        else:
            results["raw_failed"] += 1
            if results["raw_failed"] <= 3:
                # Log full error + payload keys for debugging
                err_text = resp.text[:500] if resp.text else "no response"
                payload_keys = list(clean.keys())
                results["errors"].append(f"RAW TX row {row_count}: {resp.status_code} keys={payload_keys} err={err_text}")
        time.sleep(0.05)

    wb.close()
    log.info(f"Raw transactions: {results['raw_created']} created, {results['raw_failed']} failed out of {row_count}")
    return results

# ── PIPELINE D: TRACK IMPORT BATCH (ma_IMSImportBatches) ──

def pipeline_track_import(file_type, filename, dist_guid, dist_set, total_rows, parsed_rows, failed_rows, processing_time):
    """Create an import batch record for audit trail."""
    results = {"import_tracked": False, "errors": []}

    try:
        import_set = resolve_odata_set("ma_imsimportbatches")
    except Exception:
        return results

    import_columns, import_primary = discover_valid_columns("ma_imsimportbatches")
    if not import_columns:
        return results

    payload = {"ma_distributor@odata.bind": f"/{dist_set}({dist_guid})"}

    field_map = {
        "ma_filename": filename[:200] if filename else "unknown",
        "ma_source": "Power Automate / Vercel",
    }
    for k, v in field_map.items():
        if k in import_columns:
            payload[k] = v

    int_map = {
        "ma_totalrows": total_rows,
        "ma_parsedrows": parsed_rows,
        "ma_errorrows": failed_rows,
    }
    for k, v in int_map.items():
        if k in import_columns:
            payload[k] = v

    if "ma_importdate" in import_columns:
        payload["ma_importdate"] = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    if "ma_processingtime" in import_columns:
        payload["ma_processingtime"] = int(processing_time)

    if import_primary:
        payload[import_primary] = f"{file_type}_{filename[:50]}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

    clean = safe_payload(payload, import_columns)
    resp = http_requests.post(f"{API_URL}/{import_set}", headers=get_headers(), json=clean, timeout=30)
    if resp.status_code in (200, 201, 204):
        results["import_tracked"] = True
    else:
        results["errors"].append(f"IMPORT BATCH: {resp.status_code}")

    return results

# ── PIPELINE E: UPDATE PRODUCT ITEM CODES (ma_imsproductitemcodes) ──

def pipeline_update_item_codes(file_type, file_bytes, dist_guid, brand_guids, dist_set, prod_set, all_products):
    """Extract unique item codes from file and create/update ma_imsproductitemcodes.
    Maps distributor item codes → products dynamically."""
    results = {"itemcodes_created": 0, "itemcodes_updated": 0, "itemcodes_skipped": 0, "errors": []}

    try:
        ic_set = resolve_odata_set("ma_imsproductitemcodes")
    except Exception:
        results["errors"].append("ma_imsproductitemcodes table not accessible")
        return results

    ic_columns, ic_primary = discover_valid_columns("ma_imsproductitemcodes")
    if not ic_columns:
        results["errors"].append("Could not discover item codes columns")
        return results

    # Parse file to extract unique item code → description pairs
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if "Table" in wb.sheetnames:
        ws = wb["Table"]
    header_row, headers = detect_header_row(ws)
    col = auto_map_columns(headers)
    if "material_code" in col and "product_name" not in col:
        mc_idx = col["material_code"]
        if mc_idx + 1 < len(headers):
            col["product_name"] = mc_idx + 1

    unique_items = {}  # {item_code: product_description}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item_code = get_col_str(row, col, "material_code")
        product_desc = get_col_str(row, col, "product_name")
        if item_code and product_desc and item_code not in unique_items:
            unique_items[item_code] = product_desc
    wb.close()

    log.info(f"Item codes: {len(unique_items)} unique codes found")

    for item_code, product_desc in unique_items.items():
        # Check if this item code already exists for this distributor
        filter_str = (f"ma_itemcode eq '{item_code}'"
                      f" and _ma_distributor_value eq {dist_guid}")
        check_url = f"{API_URL}/{ic_set}?$filter={filter_str}&$top=1"
        check_resp = http_requests.get(check_url, headers=get_headers(), timeout=15)

        existing = None
        if check_resp.status_code == 200:
            records = check_resp.json().get("value", [])
            if records:
                existing = records[0]

        # Match product dynamically
        product_guid, matched_name = match_product_dynamically(product_desc, all_products)

        # Build payload
        payload = {
            "ma_Distributor@odata.bind": f"/{dist_set}({dist_guid})",
        }

        if "ma_itemcode" in ic_columns:
            payload["ma_itemcode"] = item_code[:100]
        if "ma_itemdescription" in ic_columns:
            payload["ma_itemdescription"] = product_desc[:400]
        if "ma_isactiveitemcode" in ic_columns:
            payload["ma_isactiveitemcode"] = True

        # Link to product if matched
        if product_guid:
            payload["ma_Product@odata.bind"] = f"/{prod_set}({product_guid})"
            if "ma_confidence" in ic_columns:
                payload["ma_confidence"] = 0.85
        else:
            if "ma_confidence" in ic_columns:
                payload["ma_confidence"] = 0.0

        clean = safe_payload(payload, ic_columns)

        if existing:
            # PATCH existing — update description and product link
            rec_id = None
            for k, v in existing.items():
                if k.startswith("ma_") and k.endswith("id") and k != "ma_itemcode" and v:
                    rec_id = v
                    break
            if rec_id:
                resp = http_requests.patch(f"{API_URL}/{ic_set}({rec_id})",
                    headers=get_headers(), json=clean, timeout=30)
                if resp.status_code in (200, 204):
                    results["itemcodes_updated"] += 1
                else:
                    results["errors"].append(f"ITEMCODE PATCH {item_code}: {resp.status_code}")
            else:
                results["itemcodes_skipped"] += 1
        else:
            # CREATE new
            resp = http_requests.post(f"{API_URL}/{ic_set}", headers=get_headers(), json=clean, timeout=30)
            if resp.status_code in (200, 201, 204):
                results["itemcodes_created"] += 1
            else:
                if len(results["errors"]) < 3:
                    results["errors"].append(f"ITEMCODE CREATE {item_code}: {resp.status_code} {resp.text[:200]}")

        time.sleep(0.05)

    return results

# ── MAIN PROCESSING ──

def process_file(file_type, file_bytes, distributor_hint=None):
    results = {"status": "processing", "file_type": file_type, "steps": [],
        "batches_patched": 0, "batches_created": 0,
        "metrics_created": 0, "metrics_updated": 0, "raw_created": 0,
        "itemcodes_created": 0, "itemcodes_updated": 0,
        "failed": 0, "errors": []}
    _start_time = time.time()
    try:
        stock_batches, stock_product_agg, mtd_batches, mtd_product_metrics = {}, {}, {}, {}
        if file_type == "closing_stock":
            stock_batches, stock_product_agg = parse_closing_stock(file_bytes)
            results["steps"].append(f"Parsed Closing Stock: {len(stock_batches)} batches, {len(stock_product_agg)} products")
        elif file_type == "mtd_sales":
            mtd_batches, mtd_product_metrics = parse_mtd_sales(file_bytes)
            results["steps"].append(f"Parsed MTD Sales: {len(mtd_batches)} batches, {len(mtd_product_metrics)} products")

        batch_set = resolve_odata_set(TABLE_SKU_BATCHES)
        dist_set = resolve_odata_set(TABLE_DISTRIBUTORS)
        prod_set = resolve_odata_set(TABLE_PRODUCTS)

        # Dynamic distributor lookup (from email subject or fallback)
        dist_guid, dist_name = lookup_distributor(distributor_hint)
        if not dist_guid:
            results["status"] = "error"
            results["errors"].append(f"Distributor not found: {distributor_hint}")
            return results
        results["steps"].append(f"Distributor: {dist_name}")

        # Load all products for dynamic matching
        resp = http_requests.get(f"{API_URL}/{prod_set}?$top=5000", headers=get_headers(), timeout=30)
        all_products = resp.json().get("value", []) if resp.status_code == 200 else []

        # Dynamic product matching — match by brand AND by product description
        brand_guids = {}
        # First: try known brand mappings
        for info in UNICARE_PRODUCT_MAP.values():
            guid = find_product_guid(info["brand"], all_products)
            if guid:
                brand_guids[info["brand"]] = guid

        # Second: match product descriptions from MTD data dynamically
        for brand in mtd_product_metrics.keys():
            if brand not in brand_guids:
                guid, matched_name = match_product_dynamically(brand, all_products)
                if guid:
                    brand_guids[brand] = guid
                    log.info(f"Dynamic match: '{brand}' → '{matched_name}'")

        # Third: match from batch data product names
        for batch_nb, bd in {**stock_batches, **mtd_batches}.items():
            product_name = bd.get("name") or bd.get("description", "")
            brand = bd.get("brand", "")
            if brand and brand not in brand_guids and product_name:
                guid, matched_name = match_product_dynamically(product_name, all_products)
                if guid:
                    brand_guids[brand] = guid

        results["steps"].append(f"Products matched: {len(brand_guids)} ({', '.join(brand_guids.keys())})")

        valid_columns, primary_name = discover_valid_columns(TABLE_SKU_BATCHES)

        # PIPELINE A: Enrich batches
        br = pipeline_enrich_batches(stock_batches, mtd_batches, dist_guid, brand_guids,
            batch_set, dist_set, prod_set, valid_columns, primary_name, all_products)
        results["batches_patched"] = br["patched"]
        results["batches_created"] = br["created"]
        results["failed"] += br["failed"]
        results["errors"].extend(br["errors"])
        results["steps"].append(f"Batches: {br['patched']} patched, {br['created']} created")

        # PIPELINE B: Upload product metrics (with dedup)
        mr = pipeline_upload_metrics(file_type, stock_product_agg, mtd_product_metrics,
            dist_guid, brand_guids, dist_set, prod_set)
        results["metrics_created"] = mr["metrics_created"]
        results["metrics_updated"] = mr.get("metrics_updated", 0)
        results["failed"] += mr["metrics_failed"]
        results["errors"].extend(mr["errors"])
        results["steps"].append(
            f"Metrics: {mr['metrics_created']} new, {mr.get('metrics_updated', 0)} updated")

        # PIPELINE C: Upload raw transactions (MTD only)
        rr = pipeline_upload_raw_transactions(file_type, file_bytes, dist_guid, brand_guids, dist_set, prod_set)
        results["raw_created"] = rr["raw_created"]
        results["failed"] += rr["raw_failed"]
        results["errors"].extend(rr["errors"])
        if rr["raw_created"] > 0:
            results["steps"].append(f"Raw Transactions: {rr['raw_created']} rows{' ('+str(rr.get('raw_skipped',0))+' skipped as duplicates)' if rr.get('raw_skipped') else ''}")

        # PIPELINE E: Update product item codes
        ic = pipeline_update_item_codes(file_type, file_bytes, dist_guid, brand_guids, dist_set, prod_set, all_products)
        results["itemcodes_created"] = ic["itemcodes_created"]
        results["itemcodes_updated"] = ic.get("itemcodes_updated", 0)
        results["errors"].extend(ic["errors"])
        if ic["itemcodes_created"] > 0 or ic.get("itemcodes_updated", 0) > 0:
            results["steps"].append(
                f"Item Codes: {ic['itemcodes_created']} new, {ic.get('itemcodes_updated', 0)} updated in ma_imsproductitemcodes")

        # PIPELINE D: Track import batch (last — captures totals)
        processing_time = time.time() - _start_time
        total_rows = len(stock_batches) + len(mtd_batches) + rr.get("raw_created", 0)
        ir = pipeline_track_import(file_type, results.get("filename", "unknown"),
            dist_guid, dist_set, total_rows,
            br["patched"] + br["created"] + mr["metrics_created"] + mr.get("metrics_updated", 0) + rr["raw_created"],
            results["failed"], processing_time)
        if ir["import_tracked"]:
            results["steps"].append("Import batch recorded")

        results["status"] = "success"
    except Exception as e:
        results["status"] = "error"
        results["errors"].append(str(e))
        log.exception("Processing failed")
    return results

# ── FLASK APP ──
from flask import Flask, request as flask_request, jsonify
app = Flask(__name__)

@app.route("/", methods=["GET"])
def index():
    return jsonify({"service": "NBPharma IMS - Complete Pipeline", "status": "running",
        "supported_files": {"closing_stock": "batches + stock metrics", "mtd_sales": "batches + sales metrics"}})

@app.route("/api/health", methods=["GET"])
def health():
    try:
        get_token()
        resp = http_requests.get(f"{API_URL}/WhoAmI", headers=get_headers(), timeout=10)
        connected = resp.status_code == 200
    except Exception:
        connected = False
    return jsonify({"status": "healthy" if connected else "degraded", "dataverse_connected": connected})

@app.route("/api/process", methods=["POST"])
def process():
    api_key = flask_request.headers.get("X-API-Key", "")
    if api_key != API_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 401
    if not flask_request.files and not flask_request.data:
        return jsonify({"error": "No files uploaded"}), 400

    # Extract distributor hint from email subject (sent by Power Automate)
    email_subject = flask_request.headers.get("X-Email-Subject", "")
    distributor_hint = parse_distributor_from_subject(email_subject)
    log.info(f"Email subject: '{email_subject}' → distributor: '{distributor_hint}'")

    def is_excel(fn):
        return fn.lower().endswith((".xlsx", ".xls", ".xlsm"))

    files_to_process = []
    if flask_request.files:
        for key in flask_request.files:
            f = flask_request.files[key]
            fn = f.filename or key
            if not is_excel(fn):
                continue
            ft = detect_file_type(fn)
            if ft != "unknown":
                files_to_process.append((ft, f.read(), fn))
    elif flask_request.data:
        fn = flask_request.headers.get("X-Filename", "file.xlsx")
        if not is_excel(fn):
            return jsonify({"status": "skipped", "reason": f"Not Excel: {fn}"}), 200
        ft = detect_file_type(fn)
        if ft == "unknown":
            ft = "closing_stock"
        files_to_process.append((ft, flask_request.data, fn))

    if not files_to_process:
        return jsonify({"status": "skipped", "reason": "No processable Excel files"}), 200

    all_results = []
    for ft, fb, fn in files_to_process:
        log.info(f"Processing: {fn} as {ft}")
        result = process_file(ft, fb, distributor_hint=distributor_hint)
        result["filename"] = fn
        all_results.append(result)

    tp = sum(r.get("batches_patched", 0) for r in all_results)
    tc = sum(r.get("batches_created", 0) for r in all_results)
    tm = sum(r.get("metrics_created", 0) for r in all_results)
    tu = sum(r.get("metrics_updated", 0) for r in all_results)
    tr = sum(r.get("raw_created", 0) for r in all_results)
    ti = sum(r.get("itemcodes_created", 0) + r.get("itemcodes_updated", 0) for r in all_results)
    tf = sum(r.get("failed", 0) for r in all_results)
    errs = []
    for r in all_results:
        errs.extend(r.get("errors", []))

    st = "success" if all(r["status"] == "success" for r in all_results) else "partial"
    if all(r["status"] == "error" for r in all_results):
        st = "error"

    response = {"status": st, "summary": {"files_processed": len(files_to_process),
        "batches_patched": tp, "batches_created": tc,
        "metrics_created": tm, "metrics_updated": tu,
        "raw_transactions": tr, "item_codes": ti, "failed": tf},
        "files": all_results, "errors": errs[:10]}
    return jsonify(response), 200 if st in ("success", "partial") else 500

handler = app
