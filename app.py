"""
NBPharma IMS — Web Endpoint for Automated Batch Enrichment
===========================================================
Flask app that accepts distributor Excel files via HTTP POST,
processes them, and pushes data to Dataverse.

Architecture:
  Power Automate (email trigger)
    → Downloads Excel attachments
    → POSTs them to this endpoint
    → This app processes + pushes to Dataverse
    → Returns summary JSON

Deployment: Render.com free tier (or Railway, Fly.io, etc.)

Auth: Uses a refresh token captured once via browser login.
      No Azure AD App Registration required.
      Token lasts ~90 days — re-run get_refresh_token.py to renew.

Environment variables required:
  REFRESH_TOKEN  = (from get_refresh_token.py)
  DATAVERSE_URL  = https://ppcustomprojects.crm4.dynamics.com
  API_SECRET_KEY = a-random-string-to-protect-endpoint

Requirements:
  pip install flask msal requests openpyxl gunicorn
"""

import os
import io
import json
import time
import uuid
import logging
import tempfile
from datetime import datetime
from collections import defaultdict

from flask import Flask, request, jsonify
import msal
import requests as http_requests
import openpyxl

# =============================================================================
# CONFIGURATION (from environment variables)
# =============================================================================

DATAVERSE_URL = os.environ.get("DATAVERSE_URL", "https://ppcustomprojects.crm4.dynamics.com")
API_URL = f"{DATAVERSE_URL}/api/data/v9.2"

# Auth via refresh token (no App Registration needed)
# Run get_refresh_token.py locally once to capture this
REFRESH_TOKEN = os.environ.get("REFRESH_TOKEN", "")
CLIENT_ID = "51f81489-12ee-4a9e-aaae-a2591f45987d"  # Microsoft first-party Dataverse client
AUTHORITY = "https://login.microsoftonline.com/organizations"

API_SECRET_KEY = os.environ.get("API_SECRET_KEY", "nbpharma-ims-poc-2025")

SCOPES = [f"{DATAVERSE_URL}/.default"]

TABLE_SKU_BATCHES = "ma_imsskubatches"
TABLE_DISTRIBUTORS = "ma_distributors"
TABLE_PRODUCTS = "ma_imsproducts"

# Unicare Material Code → NBPharma mapping
UNICARE_PRODUCT_MAP = {
    "U190001": {"nbp_desc": "CIMZIA 200mg/ml 2 Prefilled Syringes ME", "brand": "CIMZIA"},
    "U190003": {"nbp_desc": "Orladeyo 150mg Caps 28's US1", "brand": "ORLADEYO"},
    "U190004": {"nbp_desc": "Livmarli oral solution 30ml bottle, 1's", "brand": "LIVMARLI"},
    "U190005": {"nbp_desc": "Bimzelx Inj 160mg/ml 2 Prefilled Injectors ME1", "brand": "BIMZELX"},
}

# =============================================================================
# LOGGING
# =============================================================================
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# =============================================================================
# FLASK APP
# =============================================================================
app = Flask(__name__)


# =============================================================================
# AUTH — Client Credentials (headless, no browser)
# =============================================================================
_token_cache = {"token": None, "expires_at": 0}


def get_token():
    """Acquire token using refresh token (no App Registration needed)."""
    now = time.time()
    if _token_cache["token"] and _token_cache["expires_at"] > now + 60:
        return _token_cache["token"]

    if not REFRESH_TOKEN:
        raise Exception(
            "REFRESH_TOKEN env var not set. "
            "Run get_refresh_token.py locally and paste the token into Render."
        )

    app_client = msal.PublicClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
    )

    result = app_client.acquire_token_by_refresh_token(
        refresh_token=REFRESH_TOKEN,
        scopes=SCOPES,
    )

    if "access_token" in result:
        _token_cache["token"] = result["access_token"]
        _token_cache["expires_at"] = now + result.get("expires_in", 3600)
        log.info("✅ Token acquired via refresh token")
        return result["access_token"]

    log.error(f"❌ Auth failed: {result.get('error_description', 'Unknown')}")
    raise Exception(f"Auth failed: {result.get('error_description', 'Unknown')}")


def get_headers():
    return {
        "Authorization": f"Bearer {get_token()}",
        "Content-Type": "application/json",
        "OData-MaxVersion": "4.0",
        "OData-Version": "4.0",
        "Accept": "application/json",
    }


# =============================================================================
# DATAVERSE HELPERS
# =============================================================================
_odata_cache = {}


def resolve_odata_set(logical_name):
    if logical_name in _odata_cache:
        return _odata_cache[logical_name]
    for suffix in ["s", "es", "set", ""]:
        candidate = f"{logical_name}{suffix}"
        resp = http_requests.get(f"{API_URL}/{candidate}?$top=1", headers=get_headers(), timeout=10)
        if resp.status_code == 200:
            _odata_cache[logical_name] = candidate
            return candidate
    url = f"{API_URL}/EntityDefinitions?$filter=LogicalName eq '{logical_name}'&$select=EntitySetName"
    resp = http_requests.get(url, headers=get_headers(), timeout=15)
    if resp.status_code == 200:
        entities = resp.json().get("value", [])
        if entities:
            sn = entities[0].get("EntitySetName")
            _odata_cache[logical_name] = sn
            return sn
    return f"{logical_name}s"


def discover_valid_columns(table_name):
    url = (f"{API_URL}/EntityDefinitions(LogicalName='{table_name}')"
           f"/Attributes?$select=LogicalName,AttributeType"
           f"&$filter=AttributeType ne Microsoft.Dynamics.CRM.AttributeTypeCode'Virtual'")
    resp = http_requests.get(url, headers=get_headers(), timeout=15)
    columns = set()
    primary_name = None
    if resp.status_code == 200:
        for a in resp.json().get("value", []):
            ln = a["LogicalName"]
            if ln.startswith("ma_"):
                columns.add(ln)
                if a["AttributeType"] == "String" and not primary_name:
                    if "batchid" in ln or "name" in ln:
                        primary_name = ln
    return columns, primary_name


def safe_payload(payload, valid_columns):
    clean = {}
    for k, v in payload.items():
        if "@odata.bind" in k:
            clean[k] = v
        elif k.startswith("ma_") and k in valid_columns:
            clean[k] = v
    return clean


# =============================================================================
# FILE PARSERS
# =============================================================================

def parse_closing_stock(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    batches = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        material = str(row[0]).strip() if row[0] else None
        if not material:
            continue
        batch_nb = str(row[4]).strip() if row[4] else ""
        if not batch_nb:
            continue
        from_date = row[9]
        if isinstance(from_date, str):
            parts = from_date.split(".")
            month, year = int(parts[1]), int(parts[2])
        elif isinstance(from_date, datetime):
            month, year = from_date.month, from_date.year
        else:
            month, year = 9, 2025
        expiry = row[6]
        expiry_str = expiry.strftime("%Y-%m-%d") if isinstance(expiry, datetime) else (str(expiry)[:10] if expiry else None)
        batches[batch_nb] = {
            "material": material,
            "description": str(row[1] or ""),
            "opening": int(row[7]) if row[7] is not None else 0,
            "closing": int(row[8]) if row[8] is not None else 0,
            "expiry": expiry_str,
            "month": month,
            "year": year,
            "nbp_info": UNICARE_PRODUCT_MAP.get(material, {}),
        }
    wb.close()
    return batches


def parse_mtd_sales(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    batches = defaultdict(lambda: {"qty": 0, "value": 0, "material": "", "name": "", "brand": "", "expiry": None})
    for row in ws.iter_rows(min_row=2, values_only=True):
        batch = str(row[18] or "").strip()
        if not batch:
            continue
        batches[batch]["qty"] += (row[10] or 0)
        batches[batch]["value"] += (row[11] or 0)
        batches[batch]["material"] = str(row[7] or "")
        batches[batch]["name"] = str(row[8] or "")
        batches[batch]["brand"] = str(row[21] or "")
        if row[19] and isinstance(row[19], datetime):
            batches[batch]["expiry"] = row[19].strftime("%Y-%m-%d")
    wb.close()
    return dict(batches)


def detect_file_type(filename):
    """Detect if file is Closing Stock or MTD Sales by filename patterns."""
    fn = filename.lower()
    if "closing" in fn or "stock" in fn:
        return "closing_stock"
    elif "mtd" in fn or "sales" in fn:
        return "mtd_sales"
    return "unknown"


# =============================================================================
# CORE PROCESSING LOGIC
# =============================================================================

def process_files(files_dict):
    """
    Process uploaded files and push to Dataverse.
    files_dict: {"closing_stock": bytes, "mtd_sales": bytes}
    Returns summary dict.
    """
    results = {
        "status": "processing",
        "steps": [],
        "patched": 0,
        "created": 0,
        "failed": 0,
        "errors": [],
    }

    try:
        # Step 1: Parse files
        stock_data = {}
        mtd_data = {}

        if "closing_stock" in files_dict:
            stock_data = parse_closing_stock(files_dict["closing_stock"])
            results["steps"].append(f"Parsed {len(stock_data)} batches from Closing Stock")

        if "mtd_sales" in files_dict:
            mtd_data = parse_mtd_sales(files_dict["mtd_sales"])
            results["steps"].append(f"Parsed {len(mtd_data)} batches from MTD Sales")

        all_batch_numbers = set(stock_data.keys()) | set(mtd_data.keys())
        if not all_batch_numbers:
            results["status"] = "error"
            results["errors"].append("No batch data found in uploaded files")
            return results

        results["steps"].append(f"Combined: {len(all_batch_numbers)} unique batches")

        # Step 2: Resolve Dataverse lookups
        batch_set = resolve_odata_set(TABLE_SKU_BATCHES)
        dist_set = resolve_odata_set(TABLE_DISTRIBUTORS)
        prod_set = resolve_odata_set(TABLE_PRODUCTS)

        # Lookup Unicare
        params = {"$filter": "contains(ma_distributorname,'Unicare')", "$top": 5}
        resp = http_requests.get(f"{API_URL}/{dist_set}", headers=get_headers(), params=params, timeout=15)
        dist_guid = None
        if resp.status_code == 200 and resp.json().get("value"):
            dist_guid = resp.json()["value"][0]["ma_distributorsid"]
            results["steps"].append(f"Distributor: Unicare → {dist_guid[:8]}...")

        if not dist_guid:
            results["status"] = "error"
            results["errors"].append("Unicare distributor not found in Dataverse")
            return results

        # Lookup products
        resp = http_requests.get(f"{API_URL}/{prod_set}?$top=5000", headers=get_headers(), timeout=30)
        all_products = resp.json().get("value", []) if resp.status_code == 200 else []

        def find_product_guid(brand_name):
            brand_lower = brand_name.lower()
            for p in all_products:
                for k, v in p.items():
                    if k.startswith("ma_") and isinstance(v, str) and brand_lower in v.lower():
                        return p["ma_imsproductsid"]
            return None

        brand_guids = {}
        for mat_code, info in UNICARE_PRODUCT_MAP.items():
            guid = find_product_guid(info["brand"])
            if guid:
                brand_guids[info["brand"]] = guid

        results["steps"].append(f"Products resolved: {len(brand_guids)} brands")

        # Step 3: Discover valid columns
        valid_columns, primary_name = discover_valid_columns(TABLE_SKU_BATCHES)

        # Step 4: Find existing batches in Dataverse
        found_batches = {}
        missing_batches = []

        for batch_nb in sorted(all_batch_numbers):
            filter_str = f"ma_batchnb eq '{batch_nb}'"
            url = f"{API_URL}/{batch_set}?$filter={filter_str}&$top=100"
            resp = http_requests.get(url, headers=get_headers(), timeout=15)
            if resp.status_code == 200:
                records = resp.json().get("value", [])
                if records:
                    found_batches[batch_nb] = records
                else:
                    missing_batches.append(batch_nb)
            time.sleep(0.05)

        results["steps"].append(f"PATCH: {len(found_batches)} existing | CREATE: {len(missing_batches)} new")

        # Step 5: Build and execute updates
        def build_payload(batch_nb):
            payload = {
                "ma_Distributor@odata.bind": f"/{dist_set}({dist_guid})",
                "ma_month": stock_data.get(batch_nb, {}).get("month", 9),
                "ma_year": stock_data.get(batch_nb, {}).get("year", 2025),
            }
            if batch_nb in stock_data:
                sd = stock_data[batch_nb]
                payload["ma_quantity"] = sd["closing"]
                payload["ma_openingqty"] = sd["opening"]
                payload["ma_closingqty"] = sd["closing"]
                nbp_info = sd.get("nbp_info", {})
                if nbp_info:
                    payload["ma_itemno"] = nbp_info["nbp_desc"]
                if sd["expiry"]:
                    payload["ma_expirydate"] = sd["expiry"]
                brand = nbp_info.get("brand", "")
                if brand and brand in brand_guids:
                    payload["ma_Product@odata.bind"] = f"/{prod_set}({brand_guids[brand]})"
            elif batch_nb in mtd_data:
                md = mtd_data[batch_nb]
                payload["ma_quantity"] = int(md["qty"])
                payload["ma_itemno"] = md["name"]
                payload["ma_month"] = 9
                payload["ma_year"] = 2025
                if md["expiry"]:
                    payload["ma_expirydate"] = md["expiry"]
                brand = md.get("brand", "")
                if brand and brand in brand_guids:
                    payload["ma_Product@odata.bind"] = f"/{prod_set}({brand_guids[brand]})"
            return payload

        # Execute PATCHes
        for batch_nb, records in found_batches.items():
            payload = safe_payload(build_payload(batch_nb), valid_columns)
            for rec in records:
                resp = http_requests.patch(
                    f"{API_URL}/{batch_set}({rec['ma_imsskubatchesid']})",
                    headers=get_headers(), json=payload, timeout=30,
                )
                if resp.status_code in (200, 204):
                    results["patched"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append(f"PATCH {batch_nb}: {resp.status_code}")
                time.sleep(0.1)

        # Execute CREATEs
        for batch_nb in missing_batches:
            payload = build_payload(batch_nb)
            payload["ma_batchnb"] = batch_nb
            brand = ""
            if batch_nb in stock_data:
                brand = stock_data[batch_nb].get("nbp_info", {}).get("brand", "")
            elif batch_nb in mtd_data:
                brand = mtd_data[batch_nb].get("brand", "")
            if primary_name:
                payload[primary_name] = f"{brand}_{batch_nb}_UNI_{payload.get('ma_year', 2025)}-{payload.get('ma_month', 9):02d}"

            payload = safe_payload(payload, valid_columns)

            resp = http_requests.post(
                f"{API_URL}/{batch_set}",
                headers=get_headers(), json=payload, timeout=30,
            )
            if resp.status_code in (200, 201, 204):
                results["created"] += 1
            else:
                results["failed"] += 1
                results["errors"].append(f"CREATE {batch_nb}: {resp.status_code} {resp.text[:200]}")
            time.sleep(0.1)

        results["status"] = "success"
        results["steps"].append(f"Done: {results['patched']} patched, {results['created']} created, {results['failed']} failed")

    except Exception as e:
        results["status"] = "error"
        results["errors"].append(str(e))
        log.exception("Processing failed")

    return results


# =============================================================================
# ROUTES
# =============================================================================

@app.route("/", methods=["GET"])
def health():
    return jsonify({
        "service": "NBPharma IMS Batch Enrichment",
        "status": "running",
        "version": "1.0.0",
        "endpoints": {
            "POST /process": "Upload Closing Stock + MTD Sales Excel files",
            "GET /health": "Health check",
        },
    })


@app.route("/health", methods=["GET"])
def health_check():
    # Verify Dataverse connectivity
    try:
        token = get_token()
        resp = http_requests.get(f"{API_URL}/WhoAmI", headers=get_headers(), timeout=10)
        connected = resp.status_code == 200
    except Exception:
        connected = False

    return jsonify({
        "status": "healthy" if connected else "degraded",
        "dataverse_connected": connected,
        "dataverse_url": DATAVERSE_URL,
    })


@app.route("/process", methods=["POST"])
def process_endpoint():
    """
    Accept Excel file(s) and process them.
    
    Usage from Power Automate:
      POST /process
      Headers: X-API-Key: <your-secret>
      Body: multipart/form-data with file(s)
      
    File detection is automatic by filename:
      - *closing*/*stock* → Closing Stock Report
      - *mtd*/*sales*    → MTD Sales Report
    """
    # Auth check
    api_key = request.headers.get("X-API-Key", "")
    if api_key != API_SECRET_KEY:
        return jsonify({"error": "Unauthorized. Provide X-API-Key header."}), 401

    if not request.files:
        return jsonify({"error": "No files uploaded. Send Excel files as multipart/form-data."}), 400

    # Collect files
    files_dict = {}
    file_info = []

    for key in request.files:
        f = request.files[key]
        filename = f.filename or key
        file_bytes = f.read()
        file_type = detect_file_type(filename)

        file_info.append({"name": filename, "type": file_type, "size": len(file_bytes)})

        if file_type == "closing_stock":
            files_dict["closing_stock"] = file_bytes
        elif file_type == "mtd_sales":
            files_dict["mtd_sales"] = file_bytes
        else:
            # Try to detect from content
            files_dict.setdefault("closing_stock", file_bytes)

    log.info(f"📩 Received {len(file_info)} files: {file_info}")

    if not files_dict:
        return jsonify({"error": "Could not detect file types. Name files with 'closing'/'stock' or 'mtd'/'sales'."}), 400

    # Process
    results = process_files(files_dict)
    results["files_received"] = file_info

    status_code = 200 if results["status"] == "success" else 500
    return jsonify(results), status_code


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
